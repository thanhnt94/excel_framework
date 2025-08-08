from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

import logging

logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s | %(module)s | %(lineno)d | %(funcName)s | %(levelname)s | %(message)s',
    )

def ex_latest_column(file_path, sheet_name=None, row=None, column=None):
    """
    Retrieves the latest column with data from a specified sheet in an Excel workbook.

    Author: NGUYEN TIEN THANH / KNT15083
    Last Updated: 2025-01-22

    Parameters:
    - file_path: str
        The path to the Excel file from which to retrieve data.
    - sheet_name: str, optional
        The name of the sheet to search for the latest column. If not provided, the active sheet is used.
    - row: int, optional
        The specific row to start checking from. If provided, column must also be specified.
    - column: int or str, optional
        The column number (or letter) to check for data. If specified as a letter, it will be converted to a number.

    Returns:
    - int
        The index of the last column with data in the specified range. 
        Returns 0 if the specified cell is empty, 1 if no columns are found, 
        or the maximum column number if no specific conditions are met.

    Raises:
    - ValueError
        If a column is specified without a corresponding row.
    """

    # Load the workbook and select the specified sheet
    wb = load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    # Get the last column with data in the sheet
    last_column = ws.max_column
    logging.debug(f"Last column in the sheet: {last_column}")

    # If there are no columns, return 1 (indicating the first column)
    if last_column == 0:
        logging.warning("No columns found in the sheet.")
        return 1

    # Check if only column is provided without a row
    if column is not None and row is None:
        raise ValueError("Row must be provided if column is specified.")

    # Convert column from letter to number if necessary
    if isinstance(column, str):
        column = column_index_from_string(column)

    # If both row and column are provided, find the latest column with data from that cell
    if row is not None and column is not None:
        cell_value = ws.cell(row=row, column=column).value
        if cell_value is None:
            logging.info(f"Cell ({row}, {column}) is empty. Return 0")
            return 0
        else:
            for col in range(column, last_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is None:
                    logging.info(f"Last column with data from cell ({row}, {column}): {col - 1}.")
                    return col - 1  # Return the column before the empty cell

    # If only row is provided, find the last column with data in that row
    if row is not None:
        for col in range(last_column, 0, -1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                logging.info(f"Last column with data in row {row}: {col}.")
                return col

    logging.info(f"Last column in the sheet: {last_column}.")
    return last_column

if __name__ == "__main__":
    file_path = r"C:\Users\KNT15083\Downloads\521\summary.xlsx"

    lastcol_1 = ex_latest_column(file_path, sheet_name=None, row=None, column=None)
    lastcol_2 = ex_latest_column(file_path, sheet_name=None, row=11, column=None)
    lastcol_3 = ex_latest_column(file_path, sheet_name=None, row=11, column=8)
    lastcol_4 = ex_latest_column(file_path, sheet_name=None, row=2, column="H")
