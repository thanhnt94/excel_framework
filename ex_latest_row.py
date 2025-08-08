from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

import logging

logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s | %(module)s | %(lineno)d | %(funcName)s | %(levelname)s | %(message)s',
    )

def ex_latest_row(file_path, sheet_name=None, row=None, column=None):
    """
    Retrieves the latest row with data from a specified sheet in an Excel workbook.

    Author: NGUYEN TIEN THANH / KNT15083
    Last Updated: 2025-01-22

    Parameters:
    - file_path: str
        The path to the Excel file from which to retrieve data.
    - sheet_name: str, optional
        The name of the sheet to search for the latest row. If not provided, the active sheet is used.
    - row: int, optional
        The specific row to start checking from. If provided, column must also be specified.
    - column: int or str, optional
        The column number (or letter) to check for data. If specified as a letter, it will be converted to a number.

    Returns:
    - int
        The index of the last row with data in the specified range. 
        Returns 0 if the specified cell is empty, 1 if no rows are found, 
        or the maximum row number if no specific conditions are met.
        
    Raises:
    - ValueError
        If a row is specified without a corresponding column.
    """

    # Load the workbook and select the specified sheet
    wb = load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    # Get the last row with data in the sheet
    last_row = ws.max_row
    logging.debug(f"Last row in the sheet: {last_row}")

    # If there are no rows, return 1 (indicating the first row)
    if last_row == 0:
        logging.warning("No rows found in the sheet.")
        return 1

    # Check if only row is provided without a column
    if row is not None and column is None:
        raise ValueError("Column must be provided if row is specified.")

    # Convert column from letter to number if necessary
    if isinstance(column, str):
        column = column_index_from_string(column)

    # If both row and column are provided, find the latest row with data from that cell
    if row is not None and column is not None:
        cell_value = ws.cell(row=row, column=column).value
        if cell_value is None:
            logging.info(f"Cell ({row}, {column}) is empty. Return 0")
            return 0
        else:
            for r in range(row, last_row + 1):
                if ws.cell(row=r, column=column).value is None:
                    logging.info(f"Last row with data from cell ({row}, {column}): {r - 1}.")
                    return r - 1

    # If only column is provided, find the last row with data in that column
    if column is not None:
        for r in range(last_row, 0, -1):
            if ws.cell(row=r, column=column).value is not None:
                logging.info(f"Last row with data in column {column}: {r}.")
                return r

    # If no specific conditions are met, return the maximum row
    logging.info(f"Last row in the sheet: {last_row}.")
    return last_row


if __name__ == "__main__":
    file_path = r"C:\Users\KNT15083\Downloads\521\summary.xlsx"
    lastrow_1 = ex_latest_row(file_path, sheet_name=None, row=None, column=None)
    lastrow_2 = ex_latest_row(file_path, sheet_name=None, row=None, column="H")
    lastrow_3 = ex_latest_row(file_path, sheet_name=None, row=2, column=8)
    lastrow_4 = ex_latest_row(file_path, sheet_name=None, row=6, column=8)

