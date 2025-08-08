import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(module)s | %(lineno)d | %(funcName)s | %(levelname)s | %(message)s',
)

import warnings
warnings.filterwarnings("ignore")

from PIL import Image
Image.MAX_IMAGE_PIXELS = None

import openpyxl
from openpyxl import load_workbook

def ex_find_cells_with_fomulas(file_path, sheet_name=None, filter_text=None, find_range=None):

    """
    Finds and retrieves coordinates of cells containing formulas from a specified sheet in an Excel workbook.

    Author: NGUYEN TIEN THANH / KNT15083
    Last Updated: 2025-01-22

    Parameters:
    - file_path: str
        The path to the Excel file from which to find cells with formulas.
    - sheet_name: str, optional
        The name of the sheet to search for formulas. If not provided, the active sheet is used.
    - filter_text: str, optional
        An optional string to filter the formulas. Only formulas containing this text will be returned.
    - find_range: str, optional
        A string representing the range of cells to search (e.g., "A1:C10"). If not provided, the entire sheet is searched.

    Returns:
    - list of tuples
        A list of tuples containing the coordinates (row, column) of cells that contain formulas.
        Returns an empty list if no formulas are found or if an error occurs during the process.

    Raises:
    - Exception
        Logs an error message if there is an issue accessing the specified workbook or sheet.
    """

    logging.debug(f"Starting to find formulas in '{file_path}'.")

    try:
        wb_openpyxl = load_workbook(file_path)
        replaced_cells = []
        logging.debug(f"Workbook loaded successfully with openpyxl from '{file_path}'.")

        # Get the specified sheet or default to active sheet if sheet_name is None
        if sheet_name is None:
            ws = wb_openpyxl.active
            logging.debug("Accessing the active sheet.")
        else:
            ws = wb_openpyxl[sheet_name]
            logging.debug(f"Accessing sheet by name: '{sheet_name}'")

    except Exception as e:
        logging.error(f"Error loading workbook '{file_path}': {e}")
        return []

    # Determine the range to search
    if find_range is None:
        cell_range = ws.iter_rows()  # Search all cells
        logging.debug("Searching all cells in the sheet.")
    else:
        cell_range = ws[find_range]  # Search within the specified range
        logging.debug(f"Searching within the range: '{find_range}'.")

    for row in cell_range:
        for cell in row:
            if cell is not None and cell.data_type == 'f':  # Check if cell has a formula
                if filter_text is None or filter_text in cell.value:
                    # Append the coordinate as a tuple (row, column)
                    replaced_cells.append((cell.row, cell.column))
                    logging.debug(f"Found formula at '{cell.coordinate}': '{cell.value}' in '{file_path}'.")

    logging.info(f"Total formulas found: {len(replaced_cells)}")
    return replaced_cells


if __name__ == "__main__":
    file_path = r"C:\Users\KNT15083\Downloads\521\summary.xlsx"
    search_text = "BODY"
    cell = ex_find_cells_with_fomulas(file_path, sheet_name=None, filter_text=None)
    print(cell)
