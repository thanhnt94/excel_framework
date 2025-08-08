import logging
logging.basicConfig(
        level=logging.DEBUG,
        format='%(module)s | %(lineno)d | %(funcName)s | %(levelname)s | %(message)s',
    )

import xlwings as xw

from openpyxl import load_workbook

from PIL import Image
Image.MAX_IMAGE_PIXELS = None

def ex_sheet_action(wb, action, new_sheet_name=None, sheet_identifier=None):
    """
    Performs specified actions on a sheet within a given Excel workbook.

    Author: NGUYEN TIEN THANH / KNT15083
    Last Updated: 2025-01-22

    Parameters:
    - wb: object
        The workbook object on which the action will be performed.
    - action: str
        The action to perform on the sheet. Possible actions include:
        - 'create_sheet': Create a new sheet with the provided name.
        - 'rename_sheet': Rename an existing sheet identified by either index or name.
        - 'move_sheet': Move a sheet to a new position before a target sheet.
        - 'delete_sheet': Delete a specified sheet.
        - 'copy_sheet': Copy a specified sheet to a new sheet with a provided name.
        - 'hide_sheet': Hide a specified sheet.
        - 'unhide_sheet': Unhide a specified sheet.
    - new_sheet_name: str, optional
        The name for the new sheet or the new name for the existing sheet. Required for creating, renaming, and copying sheets.
    - sheet_identifier: int or str, optional
        The index (1-based) or name of the sheet to act upon. Required for actions that modify existing sheets.

    Returns:
    - bool
        Returns True if the action is successful, or False if the action fails or is invalid.

    Logs:
    - Logs debug information when performing actions on the workbook.
    - Logs an error message if required parameters are missing or if an unknown action is specified.
    - Logs info messages when actions are successfully completed, such as creating, renaming, moving, deleting, copying, hiding, or unhiding sheets.
    """
    logging.debug(f"Performing action '{action}' on workbook: {wb.name}")

    try:
        if action == 'create_sheet':
            if new_sheet_name:
                wb.sheets.add(new_sheet_name)
                logging.info(f"Created new sheet: '{new_sheet_name}' in workbook '{wb.name}'.")
                return True
            else:
                logging.error("New sheet name must be provided for creating a sheet.")
                return False

        elif action == 'rename_sheet':
            if sheet_identifier and new_sheet_name:
                if isinstance(sheet_identifier, int):
                    sheet = wb.sheets[sheet_identifier]
                else:
                    sheet = wb.sheets[sheet_identifier]
                
                sheet.name = new_sheet_name
                logging.info(f"Renamed sheet '{sheet_identifier}' to '{new_sheet_name}' in workbook '{wb.name}'.")
                return True
            else:
                logging.error("Both sheet identifier and new sheet name must be provided for renaming.")
                return False

        elif action == 'move_sheet':
            if sheet_identifier and isinstance(sheet_identifier, int) and new_sheet_name in wb.sheetnames:
                sheet = wb.sheets[sheet_identifier]
                target_sheet = wb.sheets[new_sheet_name]
                sheet.api.Move(Before=target_sheet.api)  # Di chuyển sheet trước sheet mục tiêu
                logging.info(f"Moved sheet '{sheet.name}' before '{target_sheet.name}' in workbook '{wb.name}'.")
                return True
            else:
                logging.error("Valid sheet identifier and target sheet name must be provided for moving.")
                return False

        elif action == 'delete_sheet':
            if sheet_identifier:
                if isinstance(sheet_identifier, int):
                    sheet = wb.sheets[sheet_identifier]
                else:
                    sheet = wb.sheets[sheet_identifier]
                
                sheet.delete()
                logging.info(f"Deleted sheet '{sheet.name}' from workbook '{wb.name}'.")
                return True
            else:
                logging.error("Sheet identifier must be provided for deleting a sheet.")
                return False

        elif action == 'copy_sheet':
            if sheet_identifier and new_sheet_name:
                if isinstance(sheet_identifier, int):
                    sheet = wb.sheets[sheet_identifier]
                else:
                    sheet = wb.sheets[sheet_identifier]

                sheet.copy(after=wb.sheets[-1]) 
                wb.sheets[-1].name = new_sheet_name
                logging.info(f"Copied sheet '{sheet.name}' to new sheet '{new_sheet_name}' in workbook '{wb.name}'.")
                return True
            else:
                logging.error("Both sheet identifier and new sheet name must be provided for copying.")
                return False

        elif action == 'hide_sheet':
            if sheet_identifier:
                if isinstance(sheet_identifier, int):
                    sheet = wb.sheets[sheet_identifier]
                else:
                    sheet = wb.sheets[sheet_identifier]

                sheet.api.Visible = 0
                logging.info(f"Hid sheet '{sheet.name}' in workbook '{wb.name}'.")
                return True
            else:
                logging.error("Sheet identifier must be provided for hiding a sheet.")
                return False

        elif action == 'unhide_sheet':
            if sheet_identifier:
                if isinstance(sheet_identifier, int):
                    sheet = wb.sheets[sheet_identifier]
                else:
                    sheet = wb.sheets[sheet_identifier]

                sheet.api.Visible = -1
                logging.info(f"Unhid sheet '{sheet.name}' in workbook '{wb.name}'.")
                return True
            else:
                logging.error("Sheet identifier must be provided for unhiding a sheet.")
                return False

        else:
            logging.error(f"Unknown action: {action}")
            return False

    except Exception as e:
        logging.error(f"Error performing action '{action}' on workbook '{wb.name}': {e}")
        return False
    
if __name__ == "__main__":
    app = xw.App(visible=False)
    wb = app.books.open(r"C:\Users\KNT15083\Downloads\FY24_Q3_UV2小林-3殿宛_1812 _original\RN02753\検討書\TR-V2-S24023.xlsx")

    wb.close()
    app.quit()
