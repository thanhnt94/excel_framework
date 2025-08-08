import openpyxl
import logging

logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s | %(module)s | %(lineno)d | %(funcName)s | %(levelname)s | %(message)s',
    )

def ex_copy_range(source_info, target_info):
    """
    Copies data from a specified range in a source Excel sheet to a target Excel sheet.

    Author: NGUYEN TIEN THANH / KNT15083
    Last Updated: 2025-01-22

    Parameters:
    - source_info: tuple
        A tuple containing the source file path, source sheet name, and source range (e.g., ("source.xlsx", "Sheet1", "A1:B10")).
    - target_info: tuple
        A tuple containing the target file path, target sheet name, and target starting cell (e.g., ("target.xlsx", "Sheet2", "A1")).

    Returns:
    - None
        The function does not return any value. It saves the target file after copying the data.

    Logs:
    - Logs an error message if the source file or sheet cannot be found (not included in the current code).
    - Logs an info message when the data is successfully copied and the target file is saved (not included in the current code).
    """

    source_file, source_sheet_name, source_range = source_info
    target_file, target_sheet_name, target_start_cell = target_info

    source_wb = openpyxl.load_workbook(source_file)
    source_sheet = source_wb[source_sheet_name]

    try:
        target_wb = openpyxl.load_workbook(target_file)
    except FileNotFoundError:
        target_wb = openpyxl.Workbook()

    if target_sheet_name in target_wb.sheetnames:
        target_sheet = target_wb[target_sheet_name]
    else:
        target_sheet = target_wb.create_sheet(title=target_sheet_name)

    start_row, start_col = openpyxl.utils.cell.coordinate_to_tuple(target_start_cell)

    if source_range.lower() == "all":
        for i, row in enumerate(source_sheet.iter_rows(values_only=True)):
            for j, value in enumerate(row):
                target_sheet.cell(row=start_row + i, column=start_col + j, value=value)
    else:
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(source_range)
        for i in range(min_row, max_row + 1):
            for j in range(min_col, max_col + 1):
                value = source_sheet.cell(row=i, column=j).value
                target_sheet.cell(row=start_row + (i - min_row), column=start_col + (j - min_col), value=value)

    target_wb.save(target_file)

if __name__ == "__main__":

    # Ví dụ sử dụng
    source_info = ('source_file.xlsx', 'Sheet1', 'A5:D10')  # Sao chép từ A5 đến D10
    target_info = ('target_file.xlsx', 'CopiedSheet', 'E3')  # Dán bắt đầu từ E3
    ex_copy_range(source_info, target_info)

    # Ví dụ sao chép toàn bộ
    source_info_all = ('source_file.xlsx', 'Sheet1', 'all')  # Sao chép toàn bộ
    target_info_all = ('target_file.xlsx', 'CopiedSheet', 'A1')  # Dán bắt đầu từ A1
    ex_copy_range(source_info_all, target_info_all)
