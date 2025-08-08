import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(module)s | %(lineno)d | %(funcName)s | %(levelname)s | %(message)s',
)

import warnings
warnings.filterwarnings("ignore")

from PIL import Image
Image.MAX_IMAGE_PIXELS = None

import openpyxl
from openpyxl import load_workbook

def ex_find_cells_with_text(file_path, search_text, sheet_name=None, exact_match=False, find_range=None):
    """
    Searches for cells containing specified text in an Excel sheet and returns their coordinates.

    Author: NGUYEN TIEN THANH / KNT15083
    Last Updated: 2025-01-22

    Parameters:
    - file_path: str
        The path to the Excel file in which to search for text.
    - search_text: str
        The text to search for within the cells.
    - sheet_name: str, optional
        The name of the sheet to search in. If not provided, the active sheet is used.
    - exact_match: bool, optional
        If True, searches for an exact match of the search_text. Defaults to False for partial matches.
    - find_range: str, optional
        A string representing the range of cells to search (e.g., "A1:C10"). If not provided, the entire sheet is searched.

    Returns:
    - list of tuples
        A list of tuples containing the coordinates (row, column) of cells that contain the search_text.
        Returns an empty list if no matches are found or if an error occurs during the process.

    Raises:
    - Exception
        Logs an error message if there is an issue accessing the specified sheet or cells.
    """
    logging.debug("Starting the search for text in cells.")
    found_cells = []  # List to store coordinates of found cells

    try:
        # Load the workbook and select the specified sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name] if sheet_name else workbook.active

        # Determine the range to search
        if find_range:
            cell_range = sheet[find_range]  # Use the specified range
        else:
            cell_range = sheet.iter_rows()  # Iterate through all rows in the sheet

        for row in cell_range:
            for cell in row:
                cell_value = cell.value
                logging.debug(f"Checking cell: '{cell.coordinate}', Content: '{cell_value}'")

                if exact_match:
                    if cell_value == search_text:
                        found_cells.append((cell.row, cell.column))  # Append as (row, column)
                        logging.info(f"Found in cell (exact match): '{cell_value}', Coordinates: ({cell.row}, {cell.column})")
                else:
                    if search_text in str(cell_value):
                        found_cells.append((cell.row, cell.column))  # Append as (row, column)
                        logging.info(f"Found in cell (partial match): '{cell_value}', Coordinates: ({cell.row}, {cell.column})")

    except Exception as e:
        logging.error(f"Error while accessing cells in sheet '{sheet_name}': {e}")

    return found_cells  # Return the list of found cell coordinates

if __name__ == "__main__":
    file_path = r"C:\Users\KNT15083\Downloads\521\summary.xlsx"
    search_text = "BODY"
    cell = ex_find_cells_with_text(file_path, search_text, sheet_name=None, exact_match=False, find_range=None)
    print(cell)
